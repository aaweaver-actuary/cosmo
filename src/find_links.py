"""
find_links.py
"""
import openpyxl
import pyxlsb

from .is_xlsb import is_xlsb


def find_links_pyxlsb(wb):
    """
    Description
    -----------
    Find the links in the workbook.

    Parameters
    ----------
    wb : object
        The workbook object. Must be able to be read by pyxlsb.

    Returns
    -------
    links : list
        The list of links in the workbook.

    Raises
    ------
    ValueError
        If the workbook object is not able to be read by pyxlsb.

    Imports
    -------
    pyxlsb

    Examples
    --------
    >>> import pyxlsb
    >>> wb = pyxlsb.open_workbook("test.xlsb")
    >>> links = find_links_pyxlsb(wb)
    >>> links
    ['C:\\Users\\test\\test2.xlsb']
    """
    # try to read the workbook with pyxlsb
    # if the workbook is not able to be read by pyxlsb, raise a value error
    try:
        # use pxlsb to open the workbook and create the wb object
        with pyxlsb.open_workbook(wb) as wb:
            # get the list of links in the workbook
            links = [link[0] for link in wb.links]
    except:
        raise ValueError("The workbook object is not able to be read by pyxlsb.")

    # return the list of links in the workbook
    return links


def find_links_openpyxl(wb):
    """
    Find the links in the workbook.

    Parameters
    ----------
    wb : object
        The workbook object. Must be able to be read by openpyxl.

    Returns
    -------
    links : list
        The list of links in the workbook.

    Raises
    ------
    ValueError
        If the workbook object is not able to be read by openpyxl.

    Imports
    -------
    openpyxl

    Examples
    --------
    >>> import openpyxl
    >>> wb = openpyxl.load_workbook("test.xlsx")
    >>> links = find_links_openpyxl(wb)
    >>> links
    ['C:\\Users\\test\\test2.xlsx']
    """
    # try to read the workbook with openpyxl
    # if the workbook is not able to be read by openpyxl,
    # raise a value error
    try:
        wb = openpyxl.load_workbook(wb)
    except:
        raise ValueError("The workbook object is not able to be read by openpyxl.")

    # get the list of links in the workbook
    links = [link.target for link in wb._external_links]

    # return the list of links in the workbook
    return links


def find_links(wb):
    """
    Find the links in the workbook.

    Description
    -----------
    This function will find the links in the workbook.

    Parameters
    ----------
    wb : object
        The workbook object. Must be able to be read by openpyxl or pyxlsb.

    Returns
    -------
    links : list
        The list of links in the workbook.

    Imports
    -------
    openpyxl
    pyxlsb

    Raises
    ------
    ValueError
        If the file path does not end with ".xlsb", ".xlsx", ".xlsm", ".xls", or ".xltx".

    Examples
    --------
    >>> import openpyxl
    >>> wb = openpyxl.load_workbook("test.xlsx")
    >>> links = find_links(wb)
    >>> links
    ['C:\\Users\\test\\test2.xlsx']

    >>> import pyxlsb
    >>> wb = pyxlsb.open_workbook("test.xlsb")
    >>> links = find_links(wb)
    >>> links
    ['C:\\Users\\test\\test2.xlsb']
    """
    # get filepath from the workbook object
    filepath = wb.filename

    # check to make sure the file path ends with
    # ".xlsb", ".xlsx", ".xlsm", ".xls", or ".xltx"
    # if the file path does not end with
    # ".xlsb", ".xlsx", ".xlsm", ".xls", or ".xltx",
    # raise a value error
    if not filepath.endswith((".xlsb", ".xlsx", ".xlsm", ".xls", ".xltx")):
        raise ValueError(f"The file path {filepath} does not end with"
        + "'.xlsb', '.xlsx', '.xlsm', '.xls', or '.xltx'.")

    # if the file is an xlsb file, use pyxlsb to find the links
    if is_xlsb(filepath):
        return find_links_pyxlsb(filepath)
    # if the file is not an xlsb file, use openpyxl to find the links
    else:
        return find_links_openpyxl(filepath)
