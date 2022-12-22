"""
update_range.py
"""

import re
# pylance: disable=import-error
import openpyxl # pylance: disable=import-error
# pylance: disable=import-error
import pyxlsb

from .get_cells_from_range import get_cells_from_range
from .is_xlsb import is_xlsb

def update_range_openpyxl(wb, excel_range, value):
    """
    Description
    -----------
    This function updates a range of cells in
    an openpyxl workbook object.

    Parameters
    ----------
    wb : openpyxl.workbook.workbook.Workbook
        Workbook object.
    excel_range : dict
        Dictionary of sheet names and cell references.
    value : str, float, tuple, or list
        Value.

    Returns
    -------
    openpyxl.workbook.workbook.Workbook
        Workbook object.

    Raises
    ------
    ValueError
        If the value cannot be a list, tuple, or dict.
    ValueError
        If the value is a list, tuple, or dict and the length
        of the value is not the same as the number of cells
        in the range.
    ValueError
        If the wb object is not a wb object.
    ValueError
        If the wb object does not refer to a ".xlsx" file.

    Imports
    -------
    openpyxl
    re

    Examples
    --------
    >>> wb = openpyxl.Workbook()
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, "test")
    >>> wb["Sheet1"]["A1"].value
    'test'
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, 1)
    >>> wb["Sheet1"]["A1"].value
    1
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, (1, 2))
    >>> wb["Sheet1"]["A1"].value
    (1, 2)
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, [1, 2])
    >>> wb["Sheet1"]["A1"].value
    [1, 2]
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, {"test": "test"})
    ValueError: The value cannot be a list, tuple, or dictionary.
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, ["test", "test"])
    ValueError: The value cannot be a list, tuple, or dictionary.
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"},
    [("test", "test"), ("test", "test")])
    ValueError: The value cannot be a list, tuple, or dictionary.
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, [
        [("test", "test"), ("test", "test")], [("test", "test"), ("test", "test")]
        ])
    ValueError: The value cannot be a list, tuple, or dictionary.
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, [[1, 2], [3, 4]])
    ValueError: The value cannot be a list, tuple, or dictionary.
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, [[1, 2], [3, 4]])
    ValueError: The value cannot be a list, tuple, or dictionary.

    """
    # if the wb object is not a wb object, raise a value error
    if not isinstance(wb, openpyxl.workbook.workbook.Workbook):
        raise ValueError("The wb object is not a wb object.")

    # if the wb object does not have a file extension of .xlsx, .xlsm, or .xltx, raise a value error
    if not wb.properties.file_extension in [".xlsx", ".xlsm", ".xltx"]:
        raise ValueError("The wb object does not have a file extension of .xlsx, .xlsm, or .xltx.")

    # loop through the excel_range dictionary
    # for each sheet name, cell reference pair
    for sheet_name, cell in excel_range.items():
        # check that the sheet name is a string or integer
        if not isinstance(sheet_name, (str, int)):
            raise ValueError(f"The sheet name \"{sheet_name}\" is not a string or integer.")

        # if the sheet name is a string, check that the sheet name is in the wb object
        # if the sheet name is not in the wb object, raise a value error
        if isinstance(sheet_name, str):
            if not sheet_name in wb.sheetnames:
                raise ValueError(f"The sheet name \"{sheet_name}\" is not in the wb object.")

        # if the sheet name is an integer, check that the sheet name is in the wb object
        # if the sheet name is not in the wb object, raise a value error
        if isinstance(sheet_name, int):
            if not sheet_name in range(1, len(wb.sheetnames) + 1):
                raise ValueError(f"The sheet name {sheet_name} is not in the wb object.")

        # check that the cell reference is a string, tuple, or list
        if not isinstance(cell, (str, tuple, list)):
            raise ValueError(f"The cell reference {cell} is not a string, tuple, or list.")

        # if the cell reference is a string, check that the cell reference is in A1 notation
        if isinstance(cell, str):
            if not re.match(r"^[A-Z]+[0-9]+$", cell):
                is_a1_notation = False

        # if the cell reference is not in A1 notation,
        # check that the cell reference is in (row, column) notation and
        # that the cell reference is a tuple, with length 2
        if not is_a1_notation:
            if not isinstance(cell, tuple):
                is_rc_notation = False
            if not len(cell) == 2:
                is_rc_notation = False

        # if the cell reference is not in (row, column) notation, raise a value error
        if not is_rc_notation and not is_a1_notation:
            raise ValueError(f"The cell reference {cell}" +
            " is not in A1 notation or (row, column) notation.")

        # if the cell reference is a list, check that the cell reference is a list of strings or tuples
        if isinstance(cell, list):
            if not all(isinstance(x, (str, tuple)) for x in cell):
                raise ValueError(f"The cell reference {cell} is not a list of strings or tuples.")

        # if the cell reference is a list of strings or tuples,
        # check that the cell references are in A1 notation or in
        # (row, column) notation. if neither of these conditions are met,
        # raise a value error
        if isinstance(cell, list):
            if not all(re.match(r"^[A-Z]+[0-9]+$", x) or isinstance(x, tuple) for x in cell):
                raise ValueError(f"The cell reference {cell} is not a list of strings or tuples.")

        # if the cell reference is a string or tuple,
        # convert the cell reference to
        # a list of individual cell references
        if isinstance(cell, (str, tuple)):
            cell = [cell]

        # if the cell reference is a list, convert the cell reference to
        # a list of individual cell references
        if isinstance(cell, list):
            cell = [x for x in cell]

        # check that the value is a string, float, tuple, or list
        if not isinstance(value, (str, float, tuple, list)):
            raise ValueError(f"The value {value} is not a string, float, tuple, or list.")

        # if the value is a string or float, check that there is only one cell reference
        if isinstance(value, (str, float)):
            if not len(cell) == 1:
                raise ValueError(f"The value {value} is not a string or a float.")

        # if the value is a tuple, check that there is only one cell reference
        # if there is not only one cell reference, raise a value error
        if isinstance(value, tuple):
            if not len(cell) == 1:
                raise ValueError(f"The value {value} is not a string or a float.")

        # if the value is a list, check that there is only one cell reference
        # if there is not only one cell reference, raise a value error
        if isinstance(value, list):
            if not len(cell) == 1:
                raise ValueError(f"The value {value} is not a string or a float.")

        # if the value is a list, check that the value is a list of strings or floats
        # if the value is not a list of strings and/or floats, raise a value error
        if isinstance(value, list):
            if not all(isinstance(x, (str, float)) for x in value):
                raise ValueError(f"The value {value} is not a list of strings or floats.")

    # loop through the excel_range dictionary
    # for each sheet name, cell reference pair
    for sheet_name, cell in excel_range.items():
        # if the cell reference is a string, convert the cell reference to a list of tuples
        if isinstance(cell, str):
            cell = [openpyxl.utils.cell.coordinate_from_string(cell)]

        # if the cell reference is a list of strings, convert the cell reference to a list of tuples
        if isinstance(cell, list):
            cell = [openpyxl.utils.cell.coordinate_from_string(x) for x in cell]

        # if the value is a string, convert the value to a tuple
        if isinstance(value, str):
            value = (value,)

        # if the value is a list of strings, convert the value to a list of tuples
        if isinstance(value, list):
            value = [tuple(x) for x in value]

        # if the value is a tuple, convert the value to a list of tuples
        if isinstance(value, tuple):
            value = [value]

        # if the sheet name is a string, get the sheet object
        if isinstance(sheet_name, str):
            ws = wb[sheet_name]

        # if the sheet name is an integer, get the sheet object
        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name - 1]

        # if the cell reference is a tuple, update the cell
        if isinstance(cell, tuple):
            ws.cell(row=cell[0], column=cell[1]).value = value[0]

        # if the cell reference is a list of tuples, update the cells
        if isinstance(cell, list):
            for i in range(len(cell)):
                ws.cell(row=cell[i][0], column=cell[i][1]).value = value[i]

    # save the workbook
    wb.save(filename=wb.filename)

    # return the workbook object
    return wb


def update_range_pyxlsb(wb, excel_range, value):
    """Update a range of cells in a pyxlsb workbook object.

    Parameters
    ----------
    wb : pyxlsb.workbook.Workbook
        The pyxlsb workbook object to be updated.
    excel_range : dict
        The dictionary of cell references to be updated.
    value : str, float, tuple, list
        The value to be updated in the cell references.

    Returns
    -------
    pyxlsb.workbook.Workbook
        The updated pyxlsb workbook object.

    Raises
    ------
    ValueError
        If the wb object is not a pyxlsb workbook object.
    ValueError
        If the wb object does not have a file extension of .xlsb.
    ValueError
        If the value is not a string, float, tuple, or list.
    ValueError
        If the value is a string or float and there is more than one cell reference.
    ValueError
        If the value is a tuple or list and there is more than one cell reference.
    ValueError
        If the value is a list and the value is not a list of strings or floats.
    ValueError
        If the value is a string or float and there is more than one cell reference.
    ValueError
        If the value is a tuple or list and there is more than one cell reference.
    ValueError
        If the value is a list and the value is not a list of strings or floats.
    ValueError
        If the value is a list and the value is not a list of strings or floats.
    ValueError
        If the value is a list and the value is not a list of strings or floats.
    
    Examples
    --------
    >>> import pyxlsb
    >>> wb = pyxlsb.open_workbook('test.xlsb')
    >>> excel_range = {'Sheet1': 'A1'}
    >>> value = 'test'
    >>> wb = update_range_pyxlsb(wb, excel_range, value)
    """

    # check that the wb object is a pyxlsb workbook object
    # if the wb object is not a pyxlsb workbook object, raise a value error
    if not isinstance(wb, pyxlsb.workbook.Workbook):
        raise ValueError(f"The wb object {wb} is not a pyxlsb workbook object.")

    # check that the wb object has a file extension of .xlsb
    # if the wb object does not have a file extension of .xlsb, raise a value error
    if not wb.filename.endswith('.xlsb'):
        raise ValueError(f"The wb object {wb} does not have a file extension of .xlsb.")

    # check that the value is a string, float, tuple, or list
    # if the value is not a string, float, tuple, or list, raise a value error
    if (not isinstance(value, str) and
        not isinstance(value, float) and
        not isinstance(value, tuple) and
        not isinstance(value, list)):
        raise ValueError(f"The value {value} is not a string, float, tuple, or list.")

    # check that the value is a string or float and there is more than one cell reference
    # if the value is a string or float and there is more than one cell reference, raise a value error
    if (isinstance(value, str) or isinstance(value, float)) and len(excel_range) > 1:
        raise ValueError(f"The value {value} is a string or float and there is more than one cell reference.")

    # check that the value is a tuple or list and there is more than one cell reference
    # if the value is a tuple or list and there is more than one cell reference, raise a value error
    if (isinstance(value, tuple) or isinstance(value, list)) and len(excel_range) > 1:
        raise ValueError(f"The value {value} is a tuple or list and there is more than one cell reference.")

    # check that the value is a list and the value is not a list of strings or floats
    # if the value is a list and the value is not a list of strings or floats, raise a value error
    if isinstance(value, list) and not all(isinstance(v, str) or isinstance(v, float) for v in value):
        raise ValueError(f"The value {value} is a list and the value is not a list of strings or floats.")

    # check that the value is a string or float and there is more than one cell reference
    # if the value is a string or float and there is more than one cell reference, raise a value error
    if (isinstance(value, str) or isinstance(value, float)) and len(excel_range) > 1:
        raise ValueError(f"The value {value} is a string or float and there is more than one cell reference.")

    # check that the value is a tuple or list and there is more than one cell reference
    # if the value is a tuple or list and there is more than one cell reference, raise a value error
    if (isinstance(value, tuple) or isinstance(value, list)) and len(excel_range) > 1:
        raise ValueError(f"The value {value} is a tuple or list and there is more than one cell reference.")

    # check that the value is a list and the value is not a list of strings or floats
    # if the value is a list and the value is not a list of strings or floats, raise a value error
    if isinstance(value, list) and not all(isinstance(v, str) or isinstance(v, float) for v in value):
        raise ValueError(f"The value {value} is a list and the value is not a list of strings or floats.")

    # check that the value is a list and the value is not a list of strings or floats
    # if the value is a list and the value is not a list of strings or floats, raise a value error
    if isinstance(value, list) and not all(isinstance(v, str) or isinstance(v, float) for v in value):
        raise ValueError(f"The value {value} is a list and the value is not a list of strings or floats.")

    # check that the value is a list and the value is not a list of strings or floats
    # if the value is a list and the value is not a list of strings or floats, raise a value error
    if isinstance(value, list) and not all(isinstance(v, str) or isinstance(v, float) for v in value):
        raise ValueError(f"The value {value} is a list and the value is not a list of strings or floats.")

    # loop through the excel range dictionary
    for sheet_name, cell_range in excel_range.items():

        # get the sheet object
        sheet = wb.get_sheet_by_name(sheet_name)

        # get the cell references
        cell_refs = get_cells_from_range(cell_range)

        # check that the value is a string or float
        # if the value is a string or float, update the cell references
        if isinstance(value, str) or isinstance(value, float):
            for cell_ref in cell_refs:
                sheet.update_cell(cell_ref, value)

        # check that the value is a tuple or list
        # if the value is a tuple or list, update the cell references
        if isinstance(value, tuple) or isinstance(value, list):
            for cell_ref, v in zip(cell_refs, value):
                sheet.update_cell(cell_ref, v)

    # save the workbook
    wb.save()

    # return the workbook
    return wb

def update_range(wb, excel_range, value):
    """
    Update the cells in the excel_range input with the value input.

    Parameters
    ----------
    wb : object
        The wb object.
    excel_range : dict
        The excel_range input that is a dictionary.
    value : str, tuple, or list
        The value input.

    Returns
    -------
    wb : object
        The wb object.

    Raises
    ------
    ValueError
        If the wb object is not a wb object.
    ValueError
        If the excel_range input is not a dictionary.
    ValueError
        If the value is not a string, tuple, or list.
    ValueError
        If the value is a string or float and there is more than one cell reference.
    ValueError
        If the value is a tuple or list and there is more than one cell reference.
    ValueError  
        If the value is a list and the value is not a list of strings or floats.
    
    Examples
    --------
    >>> import pyxlsb
    >>> wb = pyxlsb.open_workbook("test.xlsb")
    >>> excel_range = {"Sheet1": "A1:A3"}
    >>> value = "test"
    >>> wb = update_range(wb, excel_range, value)
    >>> wb.worksheets[0].cells[0][0].value
    'test'
    """
    # checks that the wb object input is a wb object either of these packages can use
    # if the wb object input is not a wb object either of these packages can use, raise a value error
    if not isinstance(wb, pyxlsb.Workbook) and not isinstance(wb, openpyxl.Workbook):
        raise ValueError(f"The wb object {wb} is not a wb object.")

    # checks whether should use pyxlsb or openpyxl using is_pyxlsb function
    # if is_pyxlsb is true, use the pyxlsb version, otherwise use the openpyxl version
    if is_xlsb(wb):
        return update_range_pyxlsb(wb, excel_range, value)
    else:
        return update_range_openpyxl(wb, excel_range, value)