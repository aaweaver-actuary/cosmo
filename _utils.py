import re
import openpyxl
import pyxlsb
import pandas as pd
import numpy as np

# function taking wb object as input and returning True if 
# the workbook file extension ends with ".xlsb" and False if not
# if the wb object is not a workbook object, a value error is raised
# this funciton starts with an extremely detailed docstring, 
# including a description of the function, the input parameters, and the output, 
# value errors, and a few examples of the function in use
def is_xlsb(wb):
    """This function takes a workbook object as input and returns True if the workbook file extension ends with ".xlsb"
    and False if not.

    Parameters
    ----------
    wb : openpyxl.Workbook or pyxlsb.Workbook
        Workbook object to be checked.

    Returns
    -------
    bool
        True if the workbook file extension ends with ".xlsb" and False if not.

    Raises
    ------
    ValueError
        If the wb is not a workbook object.

    Examples
    --------
    >>> is_xlsb(wb)
    True
    """
    # test if the object passed is a workbook object, and if not, raise a value error
    if not isinstance(wb, openpyxl.Workbook) and not isinstance(wb, pyxlsb.Workbook):
        raise ValueError("wb is not a workbook object")
    # otherwise, test if the workbook is an xlsb file by checking the file extension
    # if the workbook is an xlsb file, return True
    # if the workbook is not an xlsb file, return False
    else:
        # test if the workbook is an xlsb file by checking the file extension, using a regular expression
        return re.search(r"\.xlsb$", wb.path) is not None
    

# function that uses pyxlsb library, takes a wb object as input,
# and returns a dictionary of the named ranges in the workbook,
# where the keys are the names of the named ranges and
# the values are the values of the named ranges
# this funciton starts with an extremely detailed docstring, 
# including a description of the function, the input parameters, and the output, 
# value errors, and a few examples of the function in use
def get_named_ranges_pyxlsb(wb):
    """This function uses the pyxlsb library to take a workbook object as input and return a dictionary of the named
    ranges in the workbook, where the keys are the names of the named ranges and the values are the values of the named ranges.

    Parameters
    ----------
    wb : pyxlsb.Workbook
        Workbook object to be checked.

    Returns
    -------
    dict
        Dictionary of the named ranges in the workbook, where the keys are the names of the named ranges and
        the values are the values of the named ranges.

    Raises
    ------
    ValueError
        If the wb is not an xlsb file.

    Examples
    --------
    >>> get_named_ranges_pyxlsb(wb)
    {'named_range_1': 'Sheet1!$A$1:$A$2', 'named_range_2': 'Sheet1!$B$1:$B$2'}
    """
    # test if the workbook is an xlsb file
    if is_xlsb(wb):
        # return the named ranges in the workbook
        named_ranges = {}

        # loop through the named ranges in the workbook
        for name in wb.defined_names:
            # add the name and value of the named range to the dictionary
            named_ranges[name.name] = name.value
        return named_ranges
    # if the workbook is not an xlsb file, return nothing and raise a value error
    else:
        raise ValueError("wb is not an xlsb file")

# function that uses openpyxl library, takes a wb object as input, and returns a dictionary of the named ranges in the workbook, 
# where the keys are the names of the named ranges and the values are the values of the named ranges
# this funciton starts with an extremely detailed docstring, 
# including a description of the function, the input parameters, and the output, 
# value errors, and a few examples of the function in use
# returns a value error if the wb object is not an openpyxl workbook object, with file extension ".xlsx", ".xlsm", or ".xltx"
def get_named_ranges_openpyxl(wb):
    """This function uses the openpyxl library to take a workbook object as input and return a dictionary of the named
    ranges in the workbook, where the keys are the names of the named ranges and the values are the values of the named ranges.

    Parameters
    ----------
    wb : openpyxl.Workbook
        Workbook object to be checked.

    Returns
    -------
    dict
        Dictionary of the named ranges in the workbook, where the keys are the names of the named ranges and
        the values are the values of the named ranges.

    Raises
    ------
    ValueError
        If the wb is not an openpyxl workbook object, with file extension ".xlsx", ".xlsm", or ".xltx".

    Examples
    --------
    >>> get_named_ranges_openpyxl(wb)
    {'named_range_1': 'Sheet1!$A$1:$A$2', 'named_range_2': 'Sheet1!$B$1:$B$2'}
    """
    # test if the workbook is an openpyxl workbook object
    if isinstance(wb, openpyxl.Workbook):
        # test if the workbook is an openpyxl workbook object with file extension ".xlsx", ".xlsm", or ".xltx"
        if re.search(r"\.(xlsx|xlsm|xltx)$", wb.path) is not None:
            # return the named ranges in the workbook
            named_ranges = {}

            # loop through the named ranges in the workbook
            for name in wb.defined_names.definedName:
                # add the name and value of the named range to the dictionary
                named_ranges[name.name] = name.value
            return named_ranges
        # if the workbook is not an openpyxl workbook object with file extension ".xlsx", ".xlsm", or ".xltx", return nothing and raise a value error
        else:
            raise ValueError("wb is not an openpyxl workbook object, with file extension '.xlsx', '.xlsm', or '.xltx'")
    # if the workbook is not an openpyxl workbook object, return nothing and raise a value error
    else:
        raise ValueError("wb is not an openpyxl workbook object")
    

# function that takes a wb object as input and returns a dictionary of the named ranges in the workbook,
# where the keys are the names of the named ranges and the values are the values of the named ranges
# this function starts with an extremely detailed docstring, 
# including a description of the function, the input parameters,
# the output, the exceptions that can be raised, and a few examples of the function in use
# tests if the workbook is an xlsb file and uses the appropriate function to get the named ranges
# returns a value error if the object passed is not a workbook object
def get_named_ranges(wb):
    """This function takes a workbook object as input and returns a dictionary of the named ranges in the workbook,
    where the keys are the names of the named ranges and the values are the values of the named ranges.

    Parameters
    ----------
    wb : openpyxl.Workbook or pyxlsb.Workbook
        Workbook object to be checked.

    Returns
    -------
    dict
        Dictionary of the named ranges in the workbook, where the keys are the names of the named ranges and
        the values are the values of the named ranges.

    Raises
    ------
    ValueError
        If the wb is not a workbook object.

    Examples
    --------
    >>> get_named_ranges(wb)
    {'named_range_1': 'Sheet1!$A$1:$A$2', 'named_range_2': 'Sheet1!$B$1:$B$2'}
    """
    # first, test if the workbook is an xlsb file and use the appropriate function to get the named ranges
    if isinstance(wb, openpyxl.Workbook):
        return get_named_ranges_openpyxl(wb)
    # if not, test if the workbook is an xlsb file and use the appropriate function to get the named ranges
    elif isinstance(wb, pyxlsb.Workbook):
        return get_named_ranges_pyxlsb(wb)
    else:
        raise ValueError("wb is not a workbook object")


# function that takes a string representing a cell range in Excel as input and returns a list of the cells in the range
# this function starts with an extremely detailed docstring, including a description of the function, the input parameters, 
# the output, the exceptions that can be raised, and a few examples of the function in use
def get_cells_from_range(range_string):
    """This function takes a string representing a cell range in Excel as input and returns a list of the cells in the range.

    Parameters
    ----------
    range_string : str
        String representing a cell range in Excel.

    Returns
    -------
    list
        List of the cells in the range.

    Raises
    ------
    ValueError
        If the range_string is not a string.
    ValueError
        If the range_string is not a cell range in Excel.

    Examples
    --------
    >>> get_cells_from_range("A1")
    ['A1']
    >>> get_cells_from_range("A1:A2")
    ['A1', 'A2']
    >>> get_cells_from_range("A1:A2, B1:B2")
    ['A1', 'A2', 'B1', 'B2']
    """
    # check that the input is a string
    if not isinstance(range_string, str):
        raise ValueError("range_string is not a string")

    # check that the input is a cell range in Excel
    # a cell range should be 1-3 upper case letters followed by 1-7 digits, optionally separated by a colon 
    # and followed by a second cell range
    if not re.match(r"^[A-Z]{1,3}[0-9]{1,7}(:[A-Z]{1,3}[0-9]{1,7})?(,[A-Z]{1,3}[0-9]{1,7}(:[A-Z]{1,3}[0-9]{1,7})?)*$", range_string):
        raise ValueError("range_string is not a cell range in Excel")

    # split the input string on commas
    range_list = range_string.split(",")

    # initialize a list to store the cells
    cells = []

    # loop through the list of ranges
    for xlrange in range_list:
        # split xlrange on the colon
        xlrange = xlrange.split(":")

        # if the range only has one element, then it's a single cell
        if len(xlrange) == 1:
        
            # add the cell to the list of cells
            cells.append(xlrange[0])

        # if the range has two elements, then it's a range of cells
        elif len(xlrange) == 2:
        
            # get the first cell
            cell_start = xlrange[0]
            
            # get the last cell
            cell_end = xlrange[1]
            
            # get the column letter of the first cell
            column_letter_start = cell_start[0] 
            
            # get the column letter of the last cell
            column_letter_end = cell_end[0]
            
            # get the row number of the first cell
            row_number_start = int(cell_start[1:])

            # get the row number of the last cell
            row_number_end = int(cell_end[1:])            
            # loop through the columns
            for column_letter in column_letter_start + column_letter_end:

                # loop through the rows
                for row_number in range(row_number_start, row_number_end + 1):

                    #  add the cell to the list of cells
                    cells.append(column_letter + str(row_number))

        # if the range has more than two elements, then it's not a cell range in Excel
        else:
            raise ValueError("range_string is not a cell range in Excel")    
    # return the list of cells
    return cells

# function called `column_letter_from_index` that takes an integer representing a column in Excel as input and 
# returns the column letter
# this function starts with an extremely detailed docstring, 
# including a description of the function, the input parameters,
# the output, the exceptions that can be raised, and a few examples of the function in use
def column_letter_from_index(column_index):
    """This function takes an integer representing a column in Excel as input and returns the column letter.

    Parameters
    ----------
    column_index : int
        Integer representing a column in Excel.

    Returns
    -------
    str
        Column letter.

    Raises
    ------
    ValueError
        If the column_index is not an integer.
    ValueError
        If the column_index is not a column in Excel.

    Examples
    --------
    >>> column_letter_from_index(1)
    'A'
    >>> column_letter_from_index(27)
    'AA'
    >>> column_letter_from_index(702)
    'ZZ'
    """
    # check that the input is an integer
    if not isinstance(column_index, int):
        raise ValueError("column_index is not an integer")

    # check that the input is a column in Excel
    if not 1 <= column_index <= 18278:
        raise ValueError("column_index is not a column in Excel")

    # initialize a list to store the column letters
    column_letters = []

    # loop through the column index
    while column_index > 0:
        # subtract 1 from the column index
        column_index -= 1
        # get the column letter
        column_letter = chr(ord("A") + column_index % 26)
        # add the column letter to the list of column letters
        column_letters.append(column_letter)
        # divide the column index by 26
        column_index //= 26

    # return the column letter
    return "".join(column_letters[::-1])

# function called `column_index_from_string` that takes a string representing a column in Excel as input and returns the column index
# this function starts with an extremely detailed docstring, including a description of the function, the input parameters,
# the output, the exceptions that can be raised, and a few examples of the function in use
def column_index_from_string(column_string):
    """This function takes a string representing a column in Excel as input and returns the column index.

    Parameters
    ----------
    column_string : str
        String representing a column in Excel.

    Returns
    -------
    int
        Column index.

    Raises
    ------
    ValueError
        If the column_string is not a string.
    ValueError
        If the column_string is not a column in Excel.

    Examples
    --------
    >>> column_index_from_string("A")
    1
    >>> column_index_from_string("AA")
    27
    >>> column_index_from_string("AAA")
    703
    """
    # check that the input is a string
    if not isinstance(column_string, str):
        raise ValueError("column_string is not a string")

    # check that the input is a column in Excel
    # a column should be 1-3 upper case letters
    if not re.match(r"^[A-Z]{1,3}$", column_string):
        raise ValueError("column_string is not a column in Excel")

    # initialize the column index
    column_index = 0

    # loop through the characters in the string
    for i, char in enumerate(column_string):
        # add the column index for the character to the column index
        column_index += (ord(char) - 64) * 26 ** (len(column_string) - i - 1)

    # return the column index
    return column_index

# funciton that takes a list of cells as input and returns a list of the cells in R1C1 notation as tuples of the form (row, column)
# this function starts with an extremely detailed docstring, including a description of the function, the input parameters,
# the output, the exceptions that can be raised, and a few examples of the function in use
def get_cells_r1c1(cells):
    """This function takes a list of cells as input and returns a list of 
    the cells in R1C1 notation as tuples of the form (row, column).

    Parameters
    ----------
    cells : list
        List of cells.

    Returns
    -------
    list
        List of the cells in R1C1 notation as tuples of the form (row, column).

    Raises
    ------
    ValueError
        If the cells is not a list.
    ValueError
        If the cells is not a list of cells.

    Examples
    --------
    >>> get_cells_r1c1(["A1"])
    [(1, 1)]
    >>> get_cells_r1c1(["A1", "A2"])
    [(1, 1), (2, 1)]
    >>> get_cells_r1c1(["A1", "A2", "B1", "B2"])
    [(1, 1), (2, 1), (1, 2), (2, 2)]
    """
    # check that the input is a list
    if not isinstance(cells, list):
        raise ValueError("cells is not a list")

    # check that the input is a list of cells
    # a cell should be 1-3 upper case letters followed by 1-7 digits
    if not all([re.match(r"^[A-Z]{1,3}[0-9]{1,7}$", cell) for cell in cells]):
        raise ValueError("cells is not a list of cells")

    # initialize a list to store the cells in R1C1 notation
    cells_r1c1 = []

    # loop through the cells
    for cell in cells:
        # get the column letter
        column_letter = cell[0]
        # get the row number
        row_number = int(cell[1:])
        # get the column number
        column_number = column_index_from_string(column_letter)
        # add the cell in R1C1 notation to the list of cells in R1C1 notation
        cells_r1c1.append((row_number, column_number))
    
    # return the list of cells in R1C1 notation
    return cells_r1c1

# funciton that takes a list of cells  in R1C1 notation as tuples of the form (row, column) 
# as input and returns a list of the cells in A1 notation as strings
# this function starts with an extremely detailed docstring, including a description of the function, the input parameters,
# the output, the exceptions that can be raised, and a few examples of the function in use
def get_cells_a1(cells_r1c1):
    """This function takes a list of cells in R1C1 notation as tuples of the form (row, column) as input and returns a list of the cells in A1 notation as strings.

    Parameters
    ----------
    cells_r1c1 : list
        List of cells in R1C1 notation as tuples of the form (row, column).

    Returns
    -------
    list
        List of the cells in A1 notation as strings.

    Raises
    ------
    ValueError
        If the cells_r1c1 is not a list.
    ValueError
        If the cells_r1c1 is not a list of cells in R1C1 notation.

    Examples
    --------
    >>> get_cells_a1([(1, 1)])
    ["A1"]
    >>> get_cells_a1([(1, 1), (2, 1)])
    ["A1", "A2"]
    >>> get_cells_a1([(1, 1), (2, 1), (1, 2), (2, 2)])
    ["A1", "A2", "B1", "B2"]
    """
    # check that the input is a list
    if not isinstance(cells_r1c1, list):
        raise ValueError("cells_r1c1 is not a list")

    # check that the input is a list of cells in R1C1 notation
    # a cell in R1C1 notation should be a tuple of two integers
    # define a function to check that a cell is a tuple of two integers in the form (row, column)
    def cond(cell):
        return (
            # check that the cell is a tuple
            isinstance(cell, tuple) and

            # check that the cell is a tuple of two elements
            len(cell) == 2 and

            # check that the first element is an integer
            isinstance(cell[0], int) and

            # check that the second element is an integer
            isinstance(cell[1], int)
            )

    # check that the input is a list of cells in R1C1 notation
    if not all([cond(cell) for cell in cells_r1c1]):
        raise ValueError("cells_r1c1 is not a list of cells in R1C1 notation")

    # initialize a list to store the cells in A1 notation
    cells_a1 = []

    # loop through the cells in R1C1 notation
    for cell_r1c1 in cells_r1c1:
        # get the row number
        row_number = cell_r1c1[0]
        # get the column number
        column_number = cell_r1c1[1]
        # get the column letter
        column_letter = column_letter_from_index(column_number)        
        # add the cell in A1 notation to the list of cells in A1 notation
        cells_a1.append(column_letter + str(row_number))

    # return the list of cells in A1 notation
    return cells_a1

# function that takes a wb object as input, named range as input and a value as input, 
# and updates the named range with the value
# this is for a ".xlsb" file only
# this function starts with an extremely detailed docstring, including a description of the function, the input parameters,
# the output, the exceptions that can be raised, and a few examples of the function in use
# raises a value error if the named range is not found
# or if the value is not a list of the same length as the named range, 
# or if the wb object is not a wb object, or does not refer to a ".xlsb" file
# if the named range is a single cell, the value does not need to be a list
def update_named_range_pyxlsb(wb, named_range, value):
    """This function takes a wb object as input, named range as input and a value as input, 
    and updates the named range with the value. This is for a ".xlsb" file only.

    Parameters
    ----------
    wb : wb object
        wb object.
    named_range : str
        Named range.
    value : list
        Value.

    Returns
    -------
    None

    Raises
    ------
    ValueError
        If the named range is not found.
    ValueError
        If the value is not a list of the same length as the named range.
    ValueError
        If the wb object is not a wb object.
    ValueError
        If the wb object does not refer to a ".xlsb" file.

    Examples
    --------
    >>> update_named_range_xlsb(wb, "named_range", [1, 2, 3])
    None
    >>> update_named_range_xlsb(wb, "named_range", [1, 2, 3, 4])
    ValueError: The value is not a list of the same length as the named range.
    >>> update_named_range_xlsb(wb, "named_range", 1)
    None
    >>> update_named_range_xlsb(wb, "named_range", [1])
    None
    >>> update_named_range_xlsb(wb, "named_range", [1, 2, 3])
    None
    >>> update_named_range_xlsb(wb, "named_range", [1, 2, 3, 4])
    ValueError: The value is not a list of the same length as the named range.
    >>> update_named_range_xlsb(wb, "named_range", 1)
    None
    """
    # check that the wb object is a wb object
    if not isinstance(wb, wb):
        raise ValueError("The wb object is not a wb object.")

    # check that the wb object refers to a ".xlsb" file
    if not wb.filename.endswith(".xlsb"):
        raise ValueError("The wb object does not refer to a \".xlsb\" file.")

    # get the workbook object
    workbook = wb.active

    # get the named ranges
    named_ranges = workbook.get_named_ranges()

    # check that the named range is found
    if named_range not in named_ranges:
        raise ValueError("The named range is not found.")

    # get the named range
    named_range = named_ranges[named_range]

    # get the cells in the named range
    cells = named_range.destinations

    # get the cells in A1 notation
    cells_a1 = get_cells_a1(cells)

    # get the number of cells in the named range
    number_of_cells = len(cells_a1)

    # check that the value is a list of the same length as the named range
    if not isinstance(value, list):
        value = [value]
    if len(value) != number_of_cells:
        raise ValueError("The value is not a list of the same length as the named range.")

    # loop through the cells in the named range
    for i in range(number_of_cells):
        # get the cell in A1 notation
        cell_a1 = cells_a1[i]
        # get the value
        value_i = value[i]
        # update the cell with the value
        workbook[cell_a1] = value_i

# function that takes a wb object as input, named range as input and a value as input, 
# and updates the named range with the value
# this is for a openpyxl format-friendly files only
# .xlsx, .xlsm, .xltx, .xltm
# this function starts with an extremely detailed docstring, including a description of the function, the input parameters,
# the output, the exceptions that can be raised, and a few examples of the function in use
# raises a value error if the named range is not found
# or if the value is not a list of the same length as the named range, 
# or if the wb object is not a wb object, or does not refer to a file that is one 
# of .xlsx, .xlsm, .xltx, or .xltm
# if the named range is a single cell, the value does not need to be a list
def update_named_range_openpyxl(wb, named_range, value):
    """This function takes a wb object as input, named range as input and a value as input, 
    and updates the named range with the value. This is for a openpyxl format-friendly files only. 
    .xlsx, .xlsm, .xltx, .xltm

    Parameters
    ----------
    wb : wb object
        wb object.
    named_range : str
        Named range.
    value : list
        Value.

    Returns
    -------
    None

    Raises
    ------
    ValueError
        If the named range is not found.
    ValueError
        If the value is not a list of the same length as the named range.
    ValueError
        If the wb object is not a wb object.
    ValueError
        If the wb object does not refer to a file that is one of .xlsx, .xlsm, .xltx, or .xltm.

    Examples
    --------
    >>> update_named_range_xlsx(wb, "named_range", [1, 2, 3])
    None
    >>> update_named_range_xlsx(wb, "named_range", [1, 2, 3, 4])
    ValueError: The value is not a list of the same length as the named range.
    >>> update_named_range_xlsx(wb, "named_range", 1)
    None
    >>> update_named_range_xlsx(wb, "named_range", [1])
    None
    >>> update_named_range_xlsx(wb, "named_range", [1, 2, 3])
    None
    >>> update_named_range_xlsx(wb, "named_range", [1, 2, 3, 4])
    ValueError: The value is not a list of the same length as the named range.
    >>> update_named_range_xlsx(wb, "named_range", 1)
    None
    """
    # check that the wb object is a wb object
    if not isinstance(wb, wb):
        raise ValueError("The wb object is not a wb object.")

    # get the workbook object
    workbook = wb.active

    # get the named ranges
    named_ranges = workbook.defined_names.definedName

    # check that the named range is found
    if named_range not in named_ranges:
        raise ValueError("The named range is not found.")

    # get the named range
    named_range = named_ranges[named_range]

    # get the cells in the named range
    cells = named_range.destinations

    # get the cells in A1 notation
    cells_a1 = get_cells_a1(cells)

    # get the number of cells in the named range
    number_of_cells = len(cells_a1)

    # check that the value is a list of the same length as the named range
    # if the named range is a single cell, the value does not need to be a list
    if not isinstance(value, list):
        value = [value]
    if len(value) != number_of_cells:
        raise ValueError("The value is not a list of the same length as the named range.")

    # loop through the cells in the named range
    for i in range(number_of_cells):
        # get the cell in A1 notation
        cell_a1 = cells_a1[i]
        # get the value
        value_i = value[i]
        # update the cell with the value
        workbook[cell_a1] = value_i


# funciton that combines the two functions above
# takes a wb object as input, named range as input and a value as input,
# and updates the named range with the value
# first determines which of `update_named_range_openpyxl` or `update_named_range_pyxlsb` to use
# based on the file extension of the wb object
# raises a value error if the wb object is not a wb object
# or if the wb object does not refer to a file that is one of .xlsx, .xlsm, .xltx, .xltm, or .xlsb
# starts with a docstring that is extremely detailed and similar in structure to the docstrings of most of the above functions
def update_named_range(wb, named_range, value):
    """This function combines the two functions above. 
    Takes a wb object as input, named range as input and a value as input, 
    and updates the named range with the value. First determines which of `update_named_range_openpyxl` 
    or `update_named_range_pyxlsb` to use based on the file extension of the wb object.

    Parameters
    ----------
    wb : wb object
        wb object.
    named_range : str
        Named range.
    value : list
        Value.

    Returns
    -------
    None

    Raises
    ------
    ValueError
        If the wb object is not a wb object.
    ValueError
        If the wb object does not refer to a file that is one of .xlsx, .xlsm, .xltx, .xltm, or .xlsb.

    Examples
    --------
    >>> update_named_range(wb, "named_range", [1, 2, 3])
    None
    >>> update_named_range(wb, "named_range", [1, 2, 3, 4])
    ValueError: The value is not a list of the same length as the named range.
    >>> update_named_range(wb, "named_range", 1)
    None
    >>> update_named_range(wb, "named_range", [1])
    None
    >>> update_named_range(wb, "named_range", [1, 2, 3])
    None
    >>> update_named_range(wb, "named_range", [1, 2, 3, 4])
    ValueError: The value is not a list of the same length as the named range.
    >>> update_named_range(wb, "named_range", 1)
    None
    """
    # check that the wb object is a wb object
    if not isinstance(wb, wb):
        raise ValueError("The wb object is not a wb object.")

    # get the file name
    file_name = wb.file_name

    # get the file extension
    file_extension = file_name.split(".")[-1]

    # if the file extension is .xlsb, use the pyxlsb function
    if file_extension == "xlsb":
        update_named_range_pyxlsb(wb, named_range, value)
    # if the file extension is .xlsx, .xlsm, .xltx, or .xlt use the openpyxl function
    elif file_extension in ["xlsx", "xlsm", "xltx", "xltm"]:
        update_named_range_openpyxl(wb, named_range, value)
    # otherwise, raise a value error
    else:
        raise ValueError("The wb object does not refer to a file that is one of .xlsx, .xlsm, .xltx, .xltm, or .xlsb.")

# function that takes either a tuple, string, or list, and returns either a boolean or a list of booleans
# if the input is a tuple or string, returns a boolean, and if the input is a list, returns a list of booleans
# the boolean is true if the input is an excel cell reference in A1 notation, or an excel cell range reference, 
# and false otherwise
# similar docstring to the above functions
def is_a1_cell(cell):
    """This function takes either a tuple, string, or list, and returns either a boolean or a list of booleans. 
    If the input is a tuple or string, returns a boolean, and if the input is a list, returns a list of booleans. 
    The boolean is true if the input is an excel cell reference in A1 notation, or an excel cell range reference, 
    and false otherwise.

    Parameters
    ----------
    cell : tuple, str, or list
        Cell.

    Returns
    -------
    bool or list
        Boolean or list of booleans.

    Examples
    --------
    >>> is_a1_cell("A1")
    True
    >>> is_a1_cell("A1:B2")
    True
    >>> is_a1_cell("A1:B")
    False
    >>> is_a1_cell("A1B2")
    False
    >>> is_a1_cell("A1B")
    False
    >>> is_a1_cell("A1:B2", "A1:B2")
    [True, True]
    >>> is_a1_cell("A1:B2", "A1:B")
    [True, False]
    >>> is_a1_cell("A1:B2", "A1B2")
    [True, False]
    >>> is_a1_cell("A1:B2", "A1B")
    [True, False]
    >>> is_a1_cell("A1:B2", ["A1:B2", "A1:B", "A1B2", "A1B"])
    [True, False, False, False]
    """
    # if the input is a tuple or string, return a boolean
    if isinstance(cell, tuple) or isinstance(cell, str):
        # check that the cell is in A1 notation
        # if the cell is in A1 notation, return True
        # if the cell is not in A1 notation, return False
        return re.match(r"^[A-Z]+[0-9]+(:[A-Z]+[0-9]+)?$", cell) is not None
    # if the input is a list, return a list of booleans
    elif isinstance(cell, list):
        # loop through the list
        # check that each cell is in A1 notation
        # if the cell is in A1 notation, append True to the
        # list of booleans
        # if the cell is not in A1 notation, append False to the
        # list of booleans
        # return the list of booleans
        return [re.match(r"^[A-Z]+[0-9]+(:[A-Z]+[0-9]+)?$", cell) is not None for cell in cell]
    # otherwise, raise a value error
    else:
        raise ValueError("The cell is not a tuple, string, or list.")

# function that uses the above function to transform a list of excel cell references to A1 notation
# to check that a list of cells are in A1 notation or (row, column) notation,
# and returns a list where every element is a string in A1 notation
# similar docstring to the above functions, includes examples of A1 notation and (row, column) notation
def to_a1_cell(cell):
    """This function uses the above function to transform a list of excel cell references to A1 notation. 
    It checks that a list of cells are in A1 notation or (row, column) notation, 
    and returns a list where every element is a string in A1 notation.

    Parameters
    ----------
    cell : tuple, str, or list
        Cell.

    Returns
    -------
    str or list
        String or list of strings.

    Examples
    --------
    >>> to_a1_cell("A1")
    "A1"
    >>> to_a1_cell("A1:B2")
    "A1:B2"
    >>> to_a1_cell("A1:B")
    ValueError: The cell is not in A1 notation or (row, column) notation.
    >>> to_a1_cell("A1B2")
    ValueError: The cell is not in A1 notation or (row, column) notation.
    >>> to_a1_cell("A1B")
    ValueError: The cell is not in A1 notation or (row, column) notation.
    >>> to_a1_cell("A1:B2", "A1:B2")
    ["A1:B2", "A1:B2"]
    >>> to_a1_cell("A1:B2", "A1:B")
    ValueError: The cell is not in A1 notation or (row, column) notation.
    >>> to_a1_cell("A1:B2", "A1B2")
    ValueError: The cell is not in A1 notation or (row, column) notation.
    >>> to_a1_cell("A1:B2", "A1B")
    ValueError: The cell is not in A1 notation or (row, column) notation.
    >>> to_a1_cell("A1:B2", ["A1:B2", "A1:B", "A1B2", "A1B"])
    ["A1:B2", ValueError: The cell is not in A1 notation or (row, column) notation., ValueError: The cell is not in A1 notation or (row, column) notation., ValueError: The cell is not in A1 notation or (row, column) notation.]
    >>> to_a1_cell((1, 1))
    "A1"
    >>> to_a1_cell((1, 1), (1, 2))
    ["A1", "A2"]
    >>> to_a1_cell((1, 1), (1, 2), (1, 3))
    ["A1", "A2", "A3"]
    """


    # if the input is a tuple or string, return a string
    if isinstance(cell, tuple) or isinstance(cell, str):
        # check that the cell is in A1 notation
        # if the cell is in A1 notation, return the cell
        # if the cell is not in A1 notation, check that the cell is in (row, column) notation
        # if the cell is in (row, column) notation, transform the cell to A1 notation and return the cell
        # if the cell is not in (row, column) notation, raise a value error
        if re.match(r"^[A-Z]+[0-9]+(:[A-Z]+[0-9]+)?$", cell) is not None:
            return cell
        elif re.match(r"^\([0-9]+, [0-9]+\)(:\([0-9]+, [0-9]+\))?$", cell) is not None:
            return "".join([chr(int(cell[1]) + 64) + cell[3] + (":" + chr(int(cell[6]) + 64) + cell[8] if len(cell) > 9 else "") for cell in cell.split(":")])
        else:
            raise ValueError("The cell is not in A1 notation or (row, column) notation.")
    # if the input is a list, return a list of strings
    elif isinstance(cell, list):
        # loop through the list
        # check that each cell is in A1 notation
        # if the cell is in A1 notation, append the cell to the
        # list of strings
        # if the cell is not in A1 notation, check that the cell is in (row, column) notation
        # if the cell is in (row, column) notation, transform the cell to A1 notation and append the cell to the
        # list of strings
        # if the cell is not in (row, column) notation, raise a value error

        # function that uses the above function to transform a list of excel cell references to A1 notation
        def chg_to_a1(cell):
            return chr(int(cell[1]) + 64) + cell[3] + (":" + chr(int(cell[6]) + 64) + cell[8] if len(cell) > 9 else "")
        # return the list of strings

        return ["".join([ chg_to_a1(cell) for cell in cell.split(":")]) if re.match(r"^\([0-9]+, [0-9]+\)(:\([0-9]+, [0-9]+\))?$", cell) is not None else cell for cell in cell]


# function that is the opposite of `to_a1_cell`, and instead transforms to (row, column) notation
# uses get_cells_r1c1 to transform a list of excel cell references to a list of every (row, column) cell in that range
# similar docstring to the above functions
# returns a list if the input is a list, and a tuple if the input is a A1 style string or a (row, column)-style tuple
# does some testing to ensure input makes sense
def to_rc_cell(cell):
    """This function is the opposite of `to_a1_cell`, and instead transforms to (row, column) notation.
    It uses get_cells_r1c1 to transform a list of excel cell references to a list of every (row, column) cell in that range.

    Parameters
    ----------
    cell : tuple, str, or list
        Cell.

    Returns
    -------
    tuple or list
        Tuple or list of tuples.

    Examples
    --------
    >>> to_rc_cell("A1")
    (1, 1)
    >>> to_rc_cell("A1:B2")
    [(1, 1), (1, 2), (2, 1), (2, 2)]
    >>> to_rc_cell("A1:B")
    ValueError: The cell is not in A1 notation or (row, column) notation.
    >>> to_rc_cell("A1B2")
    ValueError: The cell is not in A1 notation or (row, column) notation.
    >>> to_rc_cell("A1B")
    ValueError: The cell is not in A1 notation or (row, column) notation.
    >>> to_rc_cell("A1:B2", "A1:B2")
    [(1, 1), (1, 2), (2, 1), (2, 2), (1, 1), (1, 2), (2, 1), (2, 2)]
    >>> to_rc_cell("A1:B2", "A1:B")
    ValueError: The cell is not in A1 notation or (row, column) notation.
    >>> to_rc_cell("A1:B2", "A1B2")
    ValueError: The cell is not in A1 notation or (row, column) notation.
    >>> to_rc_cell("A1:B2", "A1B")
    ValueError: The cell is not in A1 notation or (row, column) notation.
    >>> to_rc_cell("A1:B2", ["A1:B2", "A1:B", "A1B2", "A1B"])
    [[(1, 1), (1, 2), (2, 1), (2, 2)], ValueError: The cell is not in A1 notation or (row, column) notation., ValueError: The cell is not in A1 notation or (row, column) notation., ValueError: The cell is not in A1 notation or (row, column) notation.]
    """
    # if the input is a tuple or string, return a tuple
    if isinstance(cell, tuple) or isinstance(cell, str):
        # check that the cell is in A1 notation
        # if the cell is in A1 notation, return the cell
        # if the cell is not in A1 notation, check that the cell is in (row, column) notation
        # if the cell is in (row, column) notation, transform the cell to A1 notation and return the cell
        # if the cell is not in (row, column) notation, raise a value error
        if re.match(r"^[A-Z]+[0-9]+(:[A-Z]+[0-9]+)?$", cell) is not None:
            return tuple([int(cell[1]) if len(cell) == 2 else (int(cell[1:3]), int(cell[4:6])) for cell in cell.split(":")])
        elif re.match(r"^\([0-9]+, [0-9]+\)(:\([0-9]+, [0-9]+\))?$", cell) is not None:
            return tuple([int(cell[1]) if len(cell) == 3 else (int(cell[1]), int(cell[3])) for cell in cell.split(":")])
        else:
            raise ValueError("The cell is not in A1 notation or (row, column) notation.")
    # if the input is a list, return a list of tuples
    elif isinstance(cell, list):
        # loop through the list
        # check that each cell is in A1 notation
        # if the cell is in A1 notation, append the cell to the
        # list of tuples
        # if the cell is not in A1 notation, check that the cell is in (row, column) notation
        # if the cell is in (row, column) notation, transform the cell to A1 notation and append the cell to the
        # list of tuples
        # if the cell is not in (row, column) notation, raise a value error

        def get_list(cell):
            return int(cell[1]) if len(cell) == 2 else (int(cell[1:3]), int(cell[4:6]))
            
        # return the list of tuples
        return [tuple([get_list(cell) for cell in cell.split(":")]) if re.match(r"^[A-Z]+[0-9]+(:[A-Z]+[0-9]+)?$", cell) is not None else (int(cell[1]) if len(cell) == 3 else (int(cell[1]), int(cell[3]))) if re.match(r"^\([0-9]+, [0-9]+\)(:\([0-9]+, [0-9]+\))?$", cell) is not None else ValueError("The cell is not in A1 notation or (row, column) notation.") for cell in cell]


# function for updated an unnamed range of an openpyxl wb object
# similar docstring to the above functions
# takes a wb object as input, a excel_range input that is a dictionary, and a value as input,
# the dictionary keys are the sheet names and the dictionary values are the cell references
# the cells specified by these sheet name, cell pairs are updated with the value input
# the dictionary values are first converted to a list of individual cell references using the get_cells_r1c1 function
# or get_cells_a1 function as appropriate. the wb object is then updated with the value input
# raises a value error if the wb object is not a wb object, or a value error if the wb object does not have a
# file extension of .xlsx, .xlsm, or .xltx
# keys of the excel_range input are allowed to be a string for a sheet name, or an integer for a sheet index
# values of the excel_range input are allowed to be a string, tuple, or list of strings and tuples
# value input is allowed to be a string, tuple, or list of strings or tuples, depending on the cell input
# if the values of the excel_range input is a string or tuple, the value input is a string or float
# if the values of the excel_range input is a list, the value input is a list of strings and/or floats
# there must be the same number of values as there are cells in the excel_range input, or a value error is raised
# cannot make the value in a cell be a list, tuple, or dictionary, or a value error is raised
# starts with a similarly descriptive and similarly formatted docstring
def update_range_openpyxl(wb, excel_range, value):
    """This function updates a range of cells in an openpyxl workbook object.

    Parameters
    ----------
    wb : openpyxl.workbook.workbook.Workbook
        Workbook object.
    excel_range : dict
        Dictionary of sheet names and cell references.
    value : str, float, tuple, or list
        Value.

    Returns
    -------
    openpyxl.workbook.workbook.Workbook
        Workbook object.

    Examples
    --------
    >>> wb = openpyxl.Workbook()
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, "test")
    >>> wb["Sheet1"]["A1"].value
    'test'
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, 1)
    >>> wb["Sheet1"]["A1"].value
    1
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, (1, 2))
    >>> wb["Sheet1"]["A1"].value
    (1, 2)
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, [1, 2])
    >>> wb["Sheet1"]["A1"].value
    [1, 2]
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, {"test": "test"})
    ValueError: The value cannot be a list, tuple, or dictionary.
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, ["test", "test"])
    ValueError: The value cannot be a list, tuple, or dictionary.
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, [("test", "test"), ("test", "test")])
    ValueError: The value cannot be a list, tuple, or dictionary.
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, [[("test", "test"), ("test", "test")], [("test", "test"), ("test", "test")]])
    ValueError: The value cannot be a list, tuple, or dictionary.
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, [[1, 2], [3, 4]])
    ValueError: The value cannot be a list, tuple, or dictionary.
    >>> wb = update_range_openpyxl(wb, {"Sheet1": "A1"}, [[1, 2], [3, 4]])
    ValueError: The value cannot be a list, tuple, or dictionary.

    """
    # if the wb object is not a wb object, raise a value error
    if not isinstance(wb, openpyxl.workbook.workbook.Workbook):
        raise ValueError("The wb object is not a wb object.")

    # if the wb object does not have a file extension of .xlsx, .xlsm, or .xltx, raise a value error
    if not wb.properties.file_extension in [".xlsx", ".xlsm", ".xltx"]:
        raise ValueError("The wb object does not have a file extension of .xlsx, .xlsm, or .xltx.")

    # loop through the excel_range dictionary
    # for each sheet name, cell reference pair
    for sheet_name, cell in excel_range.items():
        # check that the sheet name is a string or integer
        if not isinstance(sheet_name, (str, int)):
            raise ValueError("The sheet name \"{}\" is not a string or integer.".format(sheet_name))

        # if the sheet name is a string, check that the sheet name is in the wb object
        # if the sheet name is not in the wb object, raise a value error
        if isinstance(sheet_name, str):
            if not sheet_name in wb.sheetnames:
                raise ValueError("The sheet name \"{}\" is not in the wb object.".format(sheet_name))

        # if the sheet name is an integer, check that the sheet name is in the wb object
        # if the sheet name is not in the wb object, raise a value error
        if isinstance(sheet_name, int):
            if not sheet_name in range(1, len(wb.sheetnames) + 1):
                raise ValueError("The sheet name {} is not in the wb object.".format(sheet_name))

        # check that the cell reference is a string, tuple, or list
        if not isinstance(cell, (str, tuple, list)):
            raise ValueError("The cell reference {} is not a string, tuple, or list.".format(cell))

        # if the cell reference is a string, check that the cell reference is in A1 notation
        if isinstance(cell, str):
            if not re.match(r"^[A-Z]+[0-9]+$", cell):
                is_a1_notation = False

        # if the cell reference is not in A1 notation, check that the cell reference is in (row, column) notation and 
        # that the cell reference is a tuple, with length 2
        if not is_a1_notation:
            if not isinstance(cell, tuple):
                is_rc_notation = False
            if not len(cell) == 2:
                is_rc_notation = False

        # if the cell reference is not in (row, column) notation, raise a value error
        if not is_rc_notation and not is_a1_notation:
            raise ValueError("The cell reference {} is not in A1 notation or (row, column) notation.".format(cell))

        # if the cell reference is a list, check that the cell reference is a list of strings or tuples
        if isinstance(cell, list):
            if not all(isinstance(x, (str, tuple)) for x in cell):
                raise ValueError("The cell reference {} is not a list of strings or tuples.".format(cell))

        # if the cell reference is a list of strings or tuples, check that the cell references are in A1 notation or in 
        # (row, column) notation. if neither of these conditions are met, raise a value error
        if isinstance(cell, list):
            if not all(re.match(r"^[A-Z]+[0-9]+$", x) or isinstance(x, tuple) for x in cell):
                raise ValueError("The cell reference {} is not a list of strings or tuples.".format(cell))

        # if the cell reference is a string or tuple, convert the cell reference to a list of individual cell references
        if isinstance(cell, (str, tuple)):
            cell = [cell]

        # if the cell reference is a list, convert the cell reference to a list of individual cell references
        if isinstance(cell, list):
            cell = [x for x in cell]

        # check that the value is a string, float, tuple, or list
        if not isinstance(value, (str, float, tuple, list)):
            raise ValueError("The value {} is not a string, float, tuple, or list.".format(value))

        # if the value is a string or float, check that there is only one cell reference
        if isinstance(value, (str, float)):
            if not len(cell) == 1:
                raise ValueError("The value {} is not a string or a float.".format(value))

        # if the value is a tuple, check that there is only one cell reference
        # if there is not only one cell reference, raise a value error
        if isinstance(value, tuple):
            if not len(cell) == 1:
                raise ValueError("The value {} is not a string or a float.".format(value))

        # if the value is a list, check that there is only one cell reference
        # if there is not only one cell reference, raise a value error
        if isinstance(value, list):
            if not len(cell) == 1:
                raise ValueError("The value {} is not a string or a float.".format(value))

        # if the value is a list, check that the value is a list of strings or floats
        # if the value is not a list of strings and/or floats, raise a value error
        if isinstance(value, list):
            if not all(isinstance(x, (str, float)) for x in value):
                raise ValueError("The value {} is not a list of strings or floats.".format(value))

    # loop through the excel_range dictionary
    # for each sheet name, cell reference pair
    for sheet_name, cell in excel_range.items():
        # if the cell reference is a string, convert the cell reference to a list of tuples
        if isinstance(cell, str):
            cell = [openpyxl.utils.cell.coordinate_from_string(cell)]

        # if the cell reference is a list of strings, convert the cell reference to a list of tuples
        if isinstance(cell, list):
            cell = [openpyxl.utils.cell.coordinate_from_string(x) for x in cell]

        # if the value is a string, convert the value to a tuple
        if isinstance(value, str):
            value = (value,)

        # if the value is a list of strings, convert the value to a list of tuples
        if isinstance(value, list):
            value = [tuple(x) for x in value]

        # if the value is a tuple, convert the value to a list of tuples
        if isinstance(value, tuple):
            value = [value]

        # if the sheet name is a string, get the sheet object
        if isinstance(sheet_name, str):
            ws = wb[sheet_name]

        # if the sheet name is an integer, get the sheet object
        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name - 1]

        # if the cell reference is a tuple, update the cell
        if isinstance(cell, tuple):
            ws.cell(row=cell[0], column=cell[1]).value = value[0]

        # if the cell reference is a list of tuples, update the cells
        if isinstance(cell, list):
            for i in range(len(cell)):
                ws.cell(row=cell[i][0], column=cell[i][1]).value = value[i]

    # save the workbook
    wb.save(filename)

    # return the workbook object
    return wb

# function for updating an unnamed range of an pyxlsb wb object
# replicates the functionality of the update_range_openpyxl function
# named update_range_pyxlsb
# similar docstring to the above functions
# takes a wb object as input, a excel_range input that is a dictionary, and a value as input,
# the dictionary keys are the sheet names and the dictionary values are the cell references
# the cells specified by these sheet name, cell pairs are updated with the value input
# the dictionary values are first converted to a list of individual cell references using the get_cells_r1c1 function
# or get_cells_a1 function as appropriate. the wb object is then updated with the value input
# raises a value error if the wb object is not a wb object, or a value error if the wb object does not have a
# file extension of .xlsb
# keys of the excel_range input are allowed to be a string for a sheet name, or an integer for a sheet index
# values of the excel_range input are allowed to be a string, tuple, or list of strings and tuples
# value input is allowed to be a string, tuple, or list of strings or tuples, depending on the cell input
# if the values of the excel_range input is a string or tuple, the value input is a string or float
# if the values of the excel_range input is a list, the value input is a list of strings and/or floats
# there must be the same number of values as there are cells in the excel_range input, or a value error is raised
# cannot make the value in a cell be a list, tuple, or dictionary, or a value error is raised
# starts with a similarly descriptive and similarly formatted docstring like above functions
def update_range_pyxlsb(wb, excel_range, value):
    """Update a range of cells in a pyxlsb workbook object.

    Parameters
    ----------
    wb : pyxlsb.workbook.Workbook
        The pyxlsb workbook object to be updated.
    excel_range : dict
        The dictionary of cell references to be updated.
    value : str, float, tuple, list
        The value to be updated in the cell references.

    Returns
    -------
    pyxlsb.workbook.Workbook
        The updated pyxlsb workbook object.

    Raises
    ------
    ValueError
        If the wb object is not a pyxlsb workbook object.
    ValueError
        If the wb object does not have a file extension of .xlsb.
    ValueError  
        If the value is not a string, float, tuple, or list.
    ValueError
        If the value is a string or float and there is more than one cell reference.
    ValueError
        If the value is a tuple or list and there is more than one cell reference.
    ValueError
        If the value is a list and the value is not a list of strings or floats.
    ValueError
        If the value is a string or float and there is more than one cell reference.
    ValueError
        If the value is a tuple or list and there is more than one cell reference.
    ValueError
        If the value is a list and the value is not a list of strings or floats.
    ValueError
        If the value is a list and the value is not a list of strings or floats.
    ValueError
        If the value is a list and the value is not a list of strings or floats.
    
    Examples
    --------
    >>> import pyxlsb
    >>> wb = pyxlsb.open_workbook('test.xlsb')
    >>> excel_range = {'Sheet1': 'A1'}
    >>> value = 'test'
    >>> wb = update_range_pyxlsb(wb, excel_range, value)
    """

    # check that the wb object is a pyxlsb workbook object
    # if the wb object is not a pyxlsb workbook object, raise a value error
    if not isinstance(wb, pyxlsb.workbook.Workbook):
        raise ValueError("The wb object {} is not a pyxlsb workbook object.".format(wb))

    # check that the wb object has a file extension of .xlsb
    # if the wb object does not have a file extension of .xlsb, raise a value error
    if not wb.filename.endswith('.xlsb'):
        raise ValueError("The wb object {} does not have a file extension of .xlsb.".format(wb))

    # check that the value is a string, float, tuple, or list
    # if the value is not a string, float, tuple, or list, raise a value error
    if not isinstance(value, str) and not isinstance(value, float) and not isinstance(value, tuple) and not isinstance(value, list):
        raise ValueError("The value {} is not a string, float, tuple, or list.".format(value))

    # check that the value is a string or float and there is more than one cell reference
    # if the value is a string or float and there is more than one cell reference, raise a value error
    if (isinstance(value, str) or isinstance(value, float)) and len(excel_range) > 1:
        raise ValueError("The value {} is a string or float and there is more than one cell reference.".format(value))

    # check that the value is a tuple or list and there is more than one cell reference
    # if the value is a tuple or list and there is more than one cell reference, raise a value error
    if (isinstance(value, tuple) or isinstance(value, list)) and len(excel_range) > 1:
        raise ValueError("The value {} is a tuple or list and there is more than one cell reference.".format(value))

    # check that the value is a list and the value is not a list of strings or floats
    # if the value is a list and the value is not a list of strings or floats, raise a value error
    if isinstance(value, list) and not all(isinstance(v, str) or isinstance(v, float) for v in value):
        raise ValueError("The value {} is a list and the value is not a list of strings or floats.".format(value))

    # check that the value is a string or float and there is more than one cell reference
    # if the value is a string or float and there is more than one cell reference, raise a value error
    if (isinstance(value, str) or isinstance(value, float)) and len(excel_range) > 1:
        raise ValueError("The value {} is a string or float and there is more than one cell reference.".format(value))

    # check that the value is a tuple or list and there is more than one cell reference
    # if the value is a tuple or list and there is more than one cell reference, raise a value error
    if (isinstance(value, tuple) or isinstance(value, list)) and len(excel_range) > 1:
        raise ValueError("The value {} is a tuple or list and there is more than one cell reference.".format(value))

    # check that the value is a list and the value is not a list of strings or floats
    # if the value is a list and the value is not a list of strings or floats, raise a value error
    if isinstance(value, list) and not all(isinstance(v, str) or isinstance(v, float) for v in value):
        raise ValueError("The value {} is a list and the value is not a list of strings or floats.".format(value))

    # check that the value is a list and the value is not a list of strings or floats
    # if the value is a list and the value is not a list of strings or floats, raise a value error
    if isinstance(value, list) and not all(isinstance(v, str) or isinstance(v, float) for v in value):
        raise ValueError("The value {} is a list and the value is not a list of strings or floats.".format(value))

    # check that the value is a list and the value is not a list of strings or floats
    # if the value is a list and the value is not a list of strings or floats, raise a value error
    if isinstance(value, list) and not all(isinstance(v, str) or isinstance(v, float) for v in value):
        raise ValueError("The value {} is a list and the value is not a list of strings or floats.".format(value))

    # loop through the excel range dictionary
    for sheet_name, cell_range in excel_range.items():

        # get the sheet object
        sheet = wb.get_sheet_by_name(sheet_name)

        # get the cell references
        cell_refs = get_cell_refs(cell_range)

        # check that the value is a string or float
        # if the value is a string or float, update the cell references
        if isinstance(value, str) or isinstance(value, float):
            for cell_ref in cell_refs:
                sheet.update_cell(cell_ref, value)

        # check that the value is a tuple or list
        # if the value is a tuple or list, update the cell references
        if isinstance(value, tuple) or isinstance(value, list):
            for cell_ref, v in zip(cell_refs, value):
                sheet.update_cell(cell_ref, v)

    # save the workbook
    wb.save()

    # return the workbook
    return wb

# function that combines the update_range_pyxlsb and update_range_openpyxl functions
# takes a wb object as input, a excel_range input that is a dictionary, and a value as input,
# the dictionary keys are the sheet names and the dictionary values are the cell references
# the cells specified by these sheet name, cell pairs are updated with the value input
# the dictionary values are first converted to a list of individual cell references using the get_cells_r1c1 function
# or get_cells_a1 function as appropriate. the wb object is then updated with the value input
# raises a value error if the wb object is not a wb object, or a value error if the wb object does not have a
# keys of the excel_range input are allowed to be a string for a sheet name, or an integer for a sheet index
# values of the excel_range input are allowed to be a string, tuple, or list of strings and tuples
# value input is allowed to be a string, tuple, or list of strings or tuples, depending on the cell input
# if the values of the excel_range input is a string or tuple, the value input is a string or float
# if the values of the excel_range input is a list, the value input is a list of strings and/or floats
# there must be the same number of values as there are cells in the excel_range input, or a value error is raised
# cannot make the value in a cell be a list, tuple, or dictionary, or a value error is raised
# tests if file extension requires pyxlsb or openpyxl and calls the appropriate function
# starts with a similarly descriptive and similarly formatted docstring like above functions
def update_range(wb, excel_range, value):
    """
    Update the cells in the excel_range input with the value input.

    Parameters
    ----------
    wb : object
        The wb object.
    excel_range : dict
        The excel_range input that is a dictionary.
    value : str, tuple, or list
        The value input.

    Returns
    -------
    wb : object
        The wb object.

    Raises
    ------
    ValueError
        If the wb object is not a wb object.
    ValueError
        If the excel_range input is not a dictionary.
    ValueError
        If the value is not a string, tuple, or list.
    ValueError
        If the value is a string or float and there is more than one cell reference.
    ValueError
        If the value is a tuple or list and there is more than one cell reference.
    ValueError  
        If the value is a list and the value is not a list of strings or floats.
    
    Examples
    --------
    >>> import pyxlsb
    >>> wb = pyxlsb.open_workbook("test.xlsb")
    >>> excel_range = {"Sheet1": "A1:A3"}
    >>> value = "test"
    >>> wb = update_range(wb, excel_range, value)
    >>> wb.worksheets[0].cells[0][0].value
    'test'
    """
    # checks that the wb object input is a wb object either of these packages can use
    # if the wb object input is not a wb object either of these packages can use, raise a value error
    if not isinstance(wb, pyxlsb2.Workbook) and not isinstance(wb, openpyxl.Workbook):
        raise ValueError("The wb object {} is not a wb object.".format(wb))
    
    # checks whether should use pyxlsb or openpyxl using is_pyxlsb function
    # if is_pyxlsb is true, use the pyxlsb version, otherwise use the openpyxl version
    if is_pyxlsb(wb):
        return update_range_pyxlsb(wb, excel_range, value)
    else:
        return update_range_openpyxl(wb, excel_range, value)

# function that takes an excel file and sheet name as inputs, sheet name defauts to "update_tbl"
# reads a table from the excel file and sheet name, that has sheet name, cell range, and value as columns
# converts the table to a dictionary formatted such that it that can be used to update excel files with 
# the update_range function
# takes a similar docstring as above functions
# returns the properly-formatted dictionary
def read_update_tbl(excel_file, sheet_name="update_tbl"):
    """
    Read a table from the excel file and sheet name, that has sheet name, cell range, and value as columns.

    Parameters
    ----------
    excel_file : str
        The excel file.
    sheet_name : str, optional
        The sheet name.

    Returns
    -------
    excel_range : dict
        The excel_range input that is a dictionary.

    Examples
    --------
    >>> excel_file = "test.xlsx"
    >>> sheet_name = "update_tbl"
    >>> excel_range = read_update_tbl(excel_file, sheet_name)
    >>> excel_range
    {'Sheet1': {'A1': 'test', 'A2': 'test2'}}
    """
    # read the excel file
    wb = read_excel(excel_file)

    # get the sheet object
    sheet = wb.get_sheet_by_name(sheet_name)

    # get the cell references
    cell_refs = get_cell_refs(sheet)

    # get the values
    values = get_values(sheet)

    # create the excel range dictionary
    excel_range = {}

    # loop through the cell references and values
    for cell_ref, value in zip(cell_refs, values):

        # get the sheet name
        sheet_name = cell_ref[0]

        # get the cell reference
        cell_ref = cell_ref[1]

        # check if the sheet name is in the excel range dictionary
        # if the sheet name is not in the excel range dictionary, add it
        if sheet_name not in excel_range:
            excel_range[sheet_name] = {}

        # add the cell reference and value to the excel range dictionary
        excel_range[sheet_name][cell_ref] = value

    # return the excel range dictionary
    return excel_range

# function that takes current_year as input and a n_years input that defaults to 20, and returns 
# a list of the years from (n_years - 1) years prior to the year input
# to the year input, inclusive, for a length of n_years years total
# takes a similar docstring as above functions
# returns the list of years
def get_years(current_year, n_years=20):
    """
    Get a list of the years from (n_years - 1) years prior to the year input to the year input, 
    inclusive, for a length of n_years years total.

    Parameters
    ----------
    current_year : int
        The current year.
    n_years : int, optional
        The number of years.

    Returns
    -------
    years : list
        The list of years.

    Examples
    --------
    >>> current_year = 2020
    >>> n_years = 20
    >>> years = get_years(current_year, n_years)
    >>> years
    [2001, 2002, 2003, 2004, 2005, 2006, 2007, 2008, 2009, 2010, 2011, 2012, 2013, 2014, 2015, 2016, 2017, 2018, 2019, 2020]
    """
    # get the current year
    current_year = int(current_year)

    # get the number of years
    n_years = int(n_years)

    # get the list of years
    years = list(range(current_year - n_years + 1, current_year + 1))

    # return the list of years
    return years

# function that takes a wb object from pyxlsb as input and returns a list of all the links to other excel files 
# in the wb object, including their full paths
# similar docstring as before
# returns the list of links
def get_links_pyxlsb(wb):
    """
    Get a list of all the links to other excel files in the wb object, including their full paths.

    Parameters
    ----------
    wb : object
        The wb object.

    Returns
    -------
    links : list
        The list of links.

    Raises
    ------
    ValueError
        If the wb object is not a wb object.
    
    Examples
    --------
    >>> import pyxlsb
    >>> wb = pyxlsb.open_workbook("test.xlsb")
    >>> links = get_links_pyxlsb(wb)
    >>> links
    ['C:\\Users\\test\\test2.xlsb']
    """
    # checks that the wb object input is a wb object either of these packages can use
    # if the wb object input is not a wb object either of these packages can use, raise a value error
    if not isinstance(wb, pyxlsb2.Workbook):
        raise ValueError("The wb object {} is not a wb object.".format(wb))

    # get the links
    links = wb.links

    # return the links
    return links

# function that takes a wb object from openpyxl as input and returns a list of all the links to other excel files 
# in the wb object, including their full paths
# similar docstring as before
# returns the list of links
# otherwise the same as the get_links_pyxlsb function
def get_links_openpyxl(wb):
    """
    Get a list of all the links to other excel files in the wb object, including their full paths.

    Parameters
    ----------
    wb : object
        The wb object.

    Returns
    -------
    links : list
        The list of links.

    Raises
    ------
    ValueError
        If the wb object is not a wb object.
    
    Examples
    --------
    >>> import openpyxl
    >>> wb = openpyxl.load_workbook("test.xlsx")
    >>> links = get_links_openpyxl(wb)
    >>> links
    ['C:\\Users\\test\\test2.xlsx']
    """
    # checks that the wb object input is a wb object openpyxl can use
    # if the wb object input is not a wb object openpyxl can use (an openpyxl.Workbook object), raise a value error
    if not isinstance(wb, openpyxl.Workbook):
        raise ValueError("The wb object {} is not a wb object.".format(wb))

    # checks that the wb object points to a file that ends in .xlsx, .xlsm, or .xltx
    # if the wb object does not point to a file that ends in .xlsx, .xlsm, or .xltx, raise a value error
    if not wb.path.endswith((".xlsx", ".xlsm", ".xltx")):
        raise ValueError("The wb object {} does not point to a file that ends in .xlsx, .xlsm, or .xltx.".format(wb))

    # get the links
    links = wb._external_links

    # return the links
    return links

# function that takes a wb object from pyxlsb as input and returns a list of all the links to other excel files
# in the wb object, including their full paths
# similar docstring as before
# returns the list of links
# first tests what package to use using is_pyxlsb function to test if you need to use pyxlsb or openpyxl
# based on the package used, calls the appropriate function
def get_links(wb):
    """
    Get a list of all the links to other excel files in the wb object, including their full paths.

    Parameters
    ----------
    wb : object
        The wb object.

    Returns
    -------
    links : list
        The list of links.

    Raises
    ------
    ValueError
        If the wb object is not a wb object.
    
    Examples
    --------
    >>> import pyxlsb
    >>> wb = pyxlsb.open_workbook("test.xlsb")
    >>> links = get_links(wb)
    >>> links
    ['C:\\Users\\test\\test2.xlsb']
    """
    # first test to be sure the wb object is a wb object either of these packages can use
    # if the wb object is not a wb object either of these packages can use, raise a value error
    if not is_wb(wb):
        raise ValueError("The wb object {} is not a wb object.".format(wb))
    
    # if you use pyxlsb, call the get_links_pyxlsb function
    if is_pyxlsb(wb):
        return get_links_pyxlsb(wb)
    # if you use openpyxl, call the get_links_openpyxl function
    else:
        return get_links_openpyxl(wb)

# function that takes a file path as input and extracts the quarter and year from the file name
# similar docstring as before
# returns the quarter and year
# expects the file name to have the substring number 1-4, "Q", and the year (4-digit number) with no spaces in between
# the "Q" may or may not be uppercase, so need to use re.IGNORECASE
# the year may or may not be surrounded by underscores
# there may or may not be an underscore before or after the quarter
# it could be anywhere in the file name
# the quarter and year are extracted using re.search
# the quarter may come first, or the year may come first, but either way they are separated by the "Q"
# you know it is the quarter because it is a number 1-4, you know it is the year because it is a 4-digit number
# convert both quarter and year to integers
def get_quarter_year(filepath):
    """
    Extract the quarter and year from the file name.

    Parameters
    ----------
    filepath : str
        The file path.

    Returns
    -------
    quarter : int
        The quarter.
    year : int
        The year.

    Raises
    ------
    ValueError
        If the file path does not have the substring number 1-4, "Q", and the year (4-digit number) with no spaces in between.
    
    Examples
    --------
    >>> filepath = "C:\\Users\\test\\test1Q2018.xlsb"
    >>> quarter, year = get_quarter_year(filepath)
    >>> quarter
    1
    >>> year
    2018

    >>> filepath = "C:\\Users\\test\\3Q2023 Analysis\\test_file.xlsb"
    >>> quarter, year = get_quarter_year(filepath)
    >>> quarter
    3
    >>> year
    2023

    >>> filepath = "C:\\Users\\test\\1999q2 Analysis\\test_file.xlsb"
    >>> quarter, year = get_quarter_year(filepath)
    >>> quarter
    2
    >>> year
    1999
    """
    # checks that the file path is a string
    # if the file path is not a string, raise a value error
    if not isinstance(filepath, str):
        raise ValueError("The file path {} is not a string.".format(filepath))

    # get the quarter and year
    # the quarter and year are extracted using re.search
    # the quarter may come first, or the year may come first, but either way they are separated by the "Q"

    # expects the file name to have the substring number 1-4, "Q", and the year (4-digit number) with no spaces in between
    # or the 4-digit year, "Q", and the quarter number 1-4 with no spaces in between
    # the "Q" may or may not be uppercase, so need to use re.IGNORECASE
    # you know it is the quarter because it is a number 1-4, you know it is the year because it is a 4-digit number
    
    # search for the quarter and year with quarter coming first, return None if it is not found
    quarter_year = re.search(r"([1-4])Q(\d{4})", filepath, re.IGNORECASE)
    
    # search for the quarter and year with year coming first if quarter_year is None, return None if it is not found
    quarter_year = re.search(r"(\d{4})Q([1-4])", filepath, re.IGNORECASE) if quarter_year is None else quarter_year
    
    # if both searches returned None, raise a value error
    if quarter_year is None:
        raise ValueError("The file path {} does not have the substring with a number 1-4, 'Q', and the year (4-digit number) with no spaces in between.".format(filepath))

    # otherwise convert both to integers
    quarter = int(quarter_year.group(1))
    year = int(quarter_year.group(2))

    # return the quarter and year
    return quarter, year

# function that takes a quarter and a year as input and returns the quarter and year of the previous quarter
# similar docstring as before
# returns the quarter and year of the previous quarter
# if the quarter is 1, the previous quarter is 4 of the previous year
# if the quarter is not 1, the previous quarter is the current quarter minus 1, in the current year
def prior_quarter(year, quarter):
    """
    Get the quarter and year of the previous quarter.

    Parameters
    ----------
    year : int
        The year.
    quarter : int
        The quarter.

    Returns
    -------
    prior_year : int
        The year of the previous quarter.
    prior_quarter : int
        The quarter of the previous quarter.

    Raises
    ------
    ValueError
        If the year is not an integer.
    ValueError
        If the quarter is not an integer.
    ValueError
        If the quarter is not a number 1-4.
    
    Examples
    --------
    >>> year = 2018
    >>> quarter = 1
    >>> prior_year, prior_quarter = prior_quarter(year, quarter)
    >>> prior_year
    2017
    >>> prior_quarter
    4

    >>> year = 2018
    >>> quarter = 4
    >>> prior_year, prior_quarter = prior_quarter(year, quarter)
    >>> prior_year
    2018
    >>> prior_quarter
    3
    """
    # checks that the year is an integer
    # if the year is not an integer, raise a value error
    if not isinstance(year, int):
        raise ValueError("The year {} is not an integer.".format(year))

    # checks that the quarter is an integer
    # if the quarter is not an integer, raise a value error
    if not isinstance(quarter, int):
        raise ValueError("The quarter {} is not an integer.".format(quarter))

    # checks that the quarter is a number 1-4
    # if the quarter is not a number 1-4, raise a value error
    if not 1 <= quarter <= 4:
        raise ValueError("The quarter {} is not a number 1-4.".format(quarter))

    # if the quarter is 1, the previous quarter is 4 of the previous year
    if quarter == 1:
        prior_year = year - 1
        prior_quarter = 4
    # if the quarter is not 1, the previous quarter is the current quarter minus 1, in the current year
    else:
        prior_year = year
        prior_quarter = quarter - 1

    # return the quarter and year of the previous quarter
    return prior_year, prior_quarter

# function that takes a path to an excel workbook as input and returns the wb object
# similar docstring as before
# returns the wb object
# the wb object is created differently depending on whether the file is an xlsb or an xlsx, xlsm, xls, or xltx file, 
# so first need to use is_xlsb function to determine which method to use, either pyxlsb or openpyxl
def open_workbook(filepath):
    """
    Open the workbook.

    Parameters
    ----------
    filepath : str
        The file path.

    Returns
    -------
    wb : object
        The workbook object.

    Raises
    ------
    ValueError
        If the file path does not end with ".xlsb", ".xlsx", ".xlsm", ".xls", or ".xltx".
    
    Examples
    --------
    >>> filepath = "C:\\Users\\test\\test1Q2018.xlsb"
    >>> wb = open_workbook(filepath)
    """

    # check to make sure the file path ends with ".xlsb", ".xlsx", ".xlsm", ".xls", or ".xltx"
    # if the file path does not end with ".xlsb", ".xlsx", ".xlsm", ".xls", or ".xltx", raise a value error
    if not filepath.endswith((".xlsb", ".xlsx", ".xlsm", ".xls", ".xltx")):
        raise ValueError("The file path {} does not end with '.xlsb', '.xlsx', '.xlsm', '.xls', or '.xltx'.".format(filepath))

    # if the file is an xlsb file, use pyxlsb to open the workbook
    if is_xlsb(filepath):
        with open_workbook_xlsb(filepath) as wb:
            return wb
    # if the file is not an xlsb file, use openpyxl to open the workbook
    else:
        return openpyxl.load_workbook(filepath)

# function that finds the links in a wb object
# similar docstring as before
# returns a list of the links in the workbook
# the links are found differently depending on whether the file is an xlsb or an xlsx, xlsm, xls, or xltx file,
# this version is for the xlsb files
# raise value error if wb is not able to be read by pyxlsb
def find_links_pyxlsb(wb):
    """
    Find the links in the workbook.

    Parameters
    ----------
    wb : object
        The workbook object. Must be able to be read by pyxlsb.

    Returns
    -------
    links : list
        The list of links in the workbook.

    Raises
    ------
    ValueError
        If the workbook object is not able to be read by pyxlsb.
    """
    # try to read the workbook with pyxlsb
    # if the workbook is not able to be read by pyxlsb, raise a value error
    try:
        with open_workbook_xlsb(wb) as wb:
            # get the list of links in the workbook
            links = [link[0] for link in wb.links]
    except:
        raise ValueError("The workbook object {} is not able to be read by pyxlsb.".format(wb))

    # return the list of links in the workbook
    return links

# function that finds the links in a wb object
# similar docstring as before
# returns a list of the links in the workbook
# the links are found differently depending on whether the file is an xlsb or an xlsx, xlsm, xls, or xltx file,
# this version is for the excel files openpyxl can read
# raise value error if wb is not able to be read by openpyxl
def find_links_openpyxl(wb):
    """
    Find the links in the workbook.

    Parameters
    ----------
    wb : object
        The workbook object. Must be able to be read by openpyxl.

    Returns
    -------
    links : list
        The list of links in the workbook.

    Raises
    ------
    ValueError
        If the workbook object is not able to be read by openpyxl.
    """
    # try to read the workbook with openpyxl
    # if the workbook is not able to be read by openpyxl, raise a value error
    try:
        wb = openpyxl.load_workbook(wb)
    except:
        raise ValueError("The workbook object {} is not able to be read by openpyxl.".format(wb))

    # get the list of links in the workbook
    links = [link.target for link in wb._external_links]

    # return the list of links in the workbook
    return links

# funciton that combines the find_links_pyxlsb and find_links_openpyxl functions into a single function
# similar docstring as before
# returns a list of the links in the workbook
# the links are found differently depending on whether the file is an xlsb or an xlsx, xlsm, xls, or xltx file,
# so first need to use is_xlsb function to determine which method to use, either pyxlsb or openpyxl
def find_links(filepath):
    """
    Find the links in the workbook.

    Parameters
    ----------
    filepath : str
        The file path.

    Returns
    -------
    links : list
        The list of links in the workbook.

    Raises
    ------
    ValueError
        If the file path does not end with ".xlsb", ".xlsx", ".xlsm", ".xls", or ".xltx".
    """
    # check to make sure the file path ends with ".xlsb", ".xlsx", ".xlsm", ".xls", or ".xltx"
    # if the file path does not end with ".xlsb", ".xlsx", ".xlsm", ".xls", or ".xltx", raise a value error
    if not filepath.endswith((".xlsb", ".xlsx", ".xlsm", ".xls", ".xltx")):
        raise ValueError("The file path {} does not end with '.xlsb', '.xlsx', '.xlsm', '.xls', or '.xltx'.".format(filepath))

    # if the file is an xlsb file, use pyxlsb to find the links
    if is_xlsb(filepath):
        return find_links_pyxlsb(filepath)
    # if the file is not an xlsb file, use openpyxl to find the links
    else:
        return find_links_openpyxl(filepath)


# function that takes a dictionary with keys=current links and values=desired links as input and
# updates the links in the workbook from the current links to the desired links
# similar docstring as before
# returns the wb object
# the wb object is created differently depending on whether the file is an xlsb or an xlsx, xlsm, xls, or xltx file,
# this version is for the xlsb files
# raise value error if wb is not able to be read by pyxlsb
# use find_links_pyxlsb function to find the links in the workbook
# if the links are not in the workbook, pass a message to the user and continue to the next link
# if the links are in the correct format, raise a value error
def update_links_pyxlsb(wb, links):
    """
    Update the links in the workbook from the current links to the desired links.

    Parameters
    ----------
    wb : object
        The workbook object. Must be able to be read by pyxlsb.
    links : dict
        The dictionary with the current links as keys and the desired links as values.
        Dictionary should be of the form:
            {
                current_link1: desired_link1,
                current_link2: desired_link2,
                ...
            }

    Returns
    -------
    wb : object
        The workbook object.

    Raises
    ------
    ValueError
        If the workbook object is not able to be read by pyxlsb.
    """
    # try to read the workbook with pyxlsb
    # if the workbook is not able to be read by pyxlsb, raise a value error
    try:
        with open_workbook_xlsb(wb) as wb:
            # get the list of links in the workbook
            current_links = [link[0] for link in wb.links]
    except:
        raise ValueError("The workbook object {} is not able to be read by pyxlsb.".format(wb))

    # for each link in the dictionary of links
    for link in links:
        # if the link is in the list of links in the workbook
        if link in current_links:
            # get the index of the link in the list of links in the workbook
            index = current_links.index(link)
            # update the link in the list of links in the workbook
            current_links[index] = links[link]
        # if the link is not in the list of links in the workbook
        else:
            # pass a message to the user
            print("The link {} is not in the workbook.".format(link))
            # continue to the next link
            continue

    # update the links in the workbook
    wb.links = current_links

    # return the workbook object
    return wb

# function that takes a dictionary with keys=current links and values=desired links as input and
# updates the links in the workbook from the current links to the desired links
# similar docstring as before
# returns the wb object
# the wb object is created differently depending on whether the file is an xlsb or an xlsx, xlsm, xls, or xltx file,
# this version is for the non-xlsb files
# raise value error if wb is not able to be read by openpyxl
# use find_links_openpyxl function to find the links in the workbook
# if the links are not in the workbook, pass a message to the user and continue to the next link
# if the links are in the correct format, raise a value error
def update_links_openpyxl(wb, links):
    """
    Update the links in the workbook from the current links to the desired links.

    Parameters
    ----------
    wb : object
        The workbook object. Must be able to be read by openpyxl.
    links : dict
        The dictionary with the current links as keys and the desired links as values.
        Dictionary should be of the form:
            {
                current_link1: desired_link1,
                current_link2: desired_link2,
                ...
            }

    Returns
    -------
    wb : object
        The workbook object.

    Raises
    ------
    ValueError
        If the workbook object is not able to be read by openpyxl.
    """
    # try to read the workbook with openpyxl
    # if the workbook is not able to be read by openpyxl, raise a value error
    try:
        wb = openpyxl.load_workbook(wb)
    except:
        raise ValueError("The workbook object {} is not able to be read by openpyxl.".format(wb))

    # get the list of links in the workbook
    current_links = [link.target for link in wb._external_links]

    # for each link in the dictionary of links
    for link in links:
        # if the link is in the list of links in the workbook
        if link in current_links:
            # get the index of the link in the list of links in the workbook
            index = current_links.index(link)
            # update the link in the list of links in the workbook
            current_links[index] = links[link]
        # if the link is not in the list of links in the workbook
        else:
            # pass a message to the user
            print("The link {} is not in the workbook.".format(link))
            # continue to the next link
            continue

    # update the links in the workbook
    wb._external_links = current_links

    # return the workbook object
    return wb

# function that takes a dictionary with keys=current links and values=desired links as input and
# updates the links in the workbook from the current links to the desired links
# similar docstring as before
# returns the wb object
# the wb object is created differently depending on whether the file is an xlsb or an xlsx, xlsm, xls, or xltx file,
# this version is the combination of the update_links_openpyxl and update_links_pyxlsb functions
# raise value error if wb is not able to be read by openpyxl or pyxlsb
# use find_links_openpyxl or find_links_pyxlsb function as appropriate to find the links in the workbook
# if the links are not in the workbook, pass a message to the user and continue to the next link
# if the links are in the correct format, raise a value error
def update_links(wb, links):
    """
    Update the links in the workbook from the current links to the desired links.

    Parameters
    ----------
    wb : object
        The workbook object. Must be able to be read by openpyxl or pyxlsb.
    links : dict
        The dictionary with the current links as keys and the desired links as values.
        Dictionary should be of the form:
            {
                current_link1: desired_link1,
                current_link2: desired_link2,
                ...
            }

    Returns
    -------
    wb : object
        The workbook object.

    Raises
    ------
    ValueError
        If the workbook object is not able to be read by openpyxl or pyxlsb.
    """
    # check to make sure the workbook can be read by openpyxl or pyxlsb
    # if the workbook is not able to be read by openpyxl or pyxlsb, raise a value error
    try:
        # if the workbook is an xlsb file
        if is_xlsb(wb):
            # use the find_links_pyxlsb function to find the links in the workbook
            current_links = find_links_pyxlsb(wb)
        # if the workbook is not an xlsb file
        else:
            # use the find_links_openpyxl function to find the links in the workbook
            current_links = find_links_openpyxl(wb)
    except:
        raise ValueError("The workbook object {} is not able to be read by openpyxl or pyxlsb.".format(wb))

    # if the workbook is an xlsb file
    if is_xlsb(wb):
        # use the update_links_pyxlsb function to update the links
        wb = update_links_pyxlsb(wb, links)
    # if the workbook is not an xlsb file
    else:
        # use the update_links_openpyxl function to update the links
        wb = update_links_openpyxl(wb, links)

    # return the workbook object
    return wb